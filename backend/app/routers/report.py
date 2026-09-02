from fastapi import APIRouter, Depends,HTTPException
from sqlalchemy.orm import Session
from io import BytesIO
from fastapi.responses import StreamingResponse

from reportlab.pdfgen import canvas
from openpyxl import Workbook

from app.database import get_db
from app import models, schemas

router = APIRouter(
    prefix="/reports",
    tags=["Reports"]
)


@router.post("/")
def create_report(
    report: schemas.ReportCreate,
    db: Session = Depends(get_db)
):
    new_report = models.Report(
        report_type=report.report_title,
        description=report.summary
    )

    db.add(new_report)
    db.commit()
    db.refresh(new_report)

    return new_report


@router.get("/")
def get_reports(
    db: Session = Depends(get_db)
):
    return db.query(models.Report).all()
@router.get("/project-progress/{project_id}")
def project_progress_report(
    project_id: int,
    db: Session = Depends(get_db)
):
    project = db.query(models.Project).filter(
        models.Project.id == project_id
    ).first()

    if not project:
        raise HTTPException(
            status_code=404,
            detail="Project not found"
        )

    progress_updates = db.query(
        models.ProgressUpdate
    ).filter(
        models.ProgressUpdate.project_id == project_id
    ).all()

    progress_reports = db.query(
        models.ProgressReport
    ).filter(
        models.ProgressReport.project_id == project_id
    ).all()

    delays = db.query(
        models.DelayRecord
    ).filter(
        models.DelayRecord.project_id == project_id
    ).all()

    return {
        "project_id": project.id,
        "project_name": project.name,
        "project_status": project.status,
        "progress_updates": progress_updates,
        "progress_reports": progress_reports,
        "delays": delays
    }
@router.get("/resource-utilization/{project_id}")
def resource_utilization_report(
    project_id: int,
    db: Session = Depends(get_db)
):
    project = db.query(models.Project).filter(
        models.Project.id == project_id
    ).first()

    if not project:
        raise HTTPException(
            status_code=404,
            detail="Project not found"
        )

    allocations = db.query(
        models.EquipmentAllocation
    ).filter(
        models.EquipmentAllocation.project_id == project_id
    ).all()

    return {
        "project_id": project.id,
        "project_name": project.name,
        "resource_allocations": allocations
    }
@router.get("/workforce/{project_id}")
def workforce_report(
    project_id: int,
    db: Session = Depends(get_db)
):
    project = db.query(models.Project).filter(
        models.Project.id == project_id
    ).first()

    if not project:
        raise HTTPException(
            status_code=404,
            detail="Project not found"
        )

    worker_assignments = db.query(
        models.WorkerAssignment
    ).filter(
        models.WorkerAssignment.project_id == project_id
    ).all()

    attendance = db.query(
        models.Attendance
    ).filter(
        models.Attendance.project_id == project_id
    ).all()

    payroll = db.query(
        models.Payroll
    ).filter(
        models.Payroll.project_id == project_id
    ).all()

    return {
        "project_id": project.id,
        "project_name": project.name,
        "worker_assignments": worker_assignments,
        "attendance": attendance,
        "payroll": payroll
    }
@router.get("/procurement/{project_id}")
def procurement_report(
    project_id: int,
    db: Session = Depends(get_db)
):
    project = db.query(models.Project).filter(
        models.Project.id == project_id
    ).first()

    if not project:
        raise HTTPException(
            status_code=404,
            detail="Project not found"
        )

    requests = db.query(
        models.ProcurementRequest
    ).filter(
        models.ProcurementRequest.project_id == project_id
    ).all()

    purchase_orders = db.query(
        models.PurchaseOrder
    ).filter(
        models.PurchaseOrder.project_id == project_id
    ).all()

    invoices = db.query(
        models.Invoice
    ).filter(
        models.Invoice.project_id == project_id
    ).all()

    return {
        "project_id": project.id,
        "project_name": project.name,
        "procurement_requests": requests,
        "purchase_orders": purchase_orders,
        "invoices": invoices
    }
@router.get("/budget/{project_id}")
def budget_report(
    project_id: int,
    db: Session = Depends(get_db)
):
    project = db.query(models.Project).filter(
        models.Project.id == project_id
    ).first()

    if not project:
        raise HTTPException(
            status_code=404,
            detail="Project not found"
        )

    budget = db.query(models.Budget).filter(
        models.Budget.project_id == project_id
    ).first()

    if not budget:
        raise HTTPException(
            status_code=404,
            detail="Budget not found for this project"
        )

    estimates = db.query(
        models.CostEstimate
    ).filter(
        models.CostEstimate.project_id == project_id
    ).all()

    expenses = db.query(
        models.Expense
    ).filter(
        models.Expense.project_id == project_id
    ).all()

    total_estimated = sum(
        item.estimated_amount for item in estimates
    )

    total_expenses = sum(
        item.amount for item in expenses
    )

    remaining_budget = budget.total_budget - total_expenses

    return {
        "project_id": project.id,
        "project_name": project.name,
        "total_budget": budget.total_budget,
        "total_estimated_cost": total_estimated,
        "total_actual_expenses": total_expenses,
        "remaining_budget": remaining_budget,
        "estimates": estimates,
        "expenses": expenses
    }
@router.get("/budget/{project_id}/pdf")
def budget_report_pdf(
    project_id: int,
    db: Session = Depends(get_db)
):
    project = db.query(models.Project).filter(
        models.Project.id == project_id
    ).first()

    if not project:
        raise HTTPException(
            status_code=404,
            detail="Project not found"
        )

    budget = db.query(models.Budget).filter(
        models.Budget.project_id == project_id
    ).first()

    if not budget:
        raise HTTPException(
            status_code=404,
            detail="Budget not found for this project"
        )

    expenses = db.query(models.Expense).filter(
        models.Expense.project_id == project_id
    ).all()

    total_expenses = sum(
        expense.amount for expense in expenses
    )

    remaining_budget = (
        budget.total_budget - total_expenses
    )

    buffer = BytesIO()

    pdf = canvas.Canvas(buffer)

    pdf.setTitle("Budget Report")

    pdf.drawString(50, 800, "PROJECT BUDGET REPORT")
    pdf.drawString(50, 770, f"Project: {project.name}")
    pdf.drawString(50, 750, f"Project ID: {project.id}")

    pdf.drawString(
        50, 710,
        f"Total Budget: {budget.total_budget}"
    )

    pdf.drawString(
        50, 690,
        f"Actual Expenses: {total_expenses}"
    )

    pdf.drawString(
        50, 670,
        f"Remaining Budget: {remaining_budget}"
    )

    pdf.drawString(50, 630, "Expenses:")

    y = 610

    for expense in expenses:
        pdf.drawString(
            60,
            y,
            f"{expense.category}: {expense.amount}"
        )
        y -= 20

        if y < 50:
            pdf.showPage()
            y = 800

    pdf.save()

    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition":
                f"attachment; filename=budget_report_{project_id}.pdf"
        }
    )
@router.get("/budget/{project_id}/excel")
def budget_report_excel(
    project_id: int,
    db: Session = Depends(get_db)
):
    project = db.query(models.Project).filter(
        models.Project.id == project_id
    ).first()

    if not project:
        raise HTTPException(
            status_code=404,
            detail="Project not found"
        )

    budget = db.query(models.Budget).filter(
        models.Budget.project_id == project_id
    ).first()

    if not budget:
        raise HTTPException(
            status_code=404,
            detail="Budget not found for this project"
        )

    estimates = db.query(
        models.CostEstimate
    ).filter(
        models.CostEstimate.project_id == project_id
    ).all()

    expenses = db.query(
        models.Expense
    ).filter(
        models.Expense.project_id == project_id
    ).all()

    total_estimated = sum(
        item.estimated_amount for item in estimates
    )

    total_expenses = sum(
        item.amount for item in expenses
    )

    remaining_budget = (
        budget.total_budget - total_expenses
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Budget Report"

    sheet["A1"] = "PROJECT BUDGET REPORT"
    sheet["A3"] = "Project ID"
    sheet["B3"] = project.id

    sheet["A4"] = "Project Name"
    sheet["B4"] = project.name

    sheet["A5"] = "Total Budget"
    sheet["B5"] = budget.total_budget

    sheet["A6"] = "Estimated Cost"
    sheet["B6"] = total_estimated

    sheet["A7"] = "Actual Expenses"
    sheet["B7"] = total_expenses

    sheet["A8"] = "Remaining Budget"
    sheet["B8"] = remaining_budget

    sheet["A10"] = "Expense Category"
    sheet["B10"] = "Description"
    sheet["C10"] = "Amount"

    row = 11

    for expense in expenses:
        sheet.cell(row=row, column=1).value = expense.category
        sheet.cell(row=row, column=2).value = expense.description
        sheet.cell(row=row, column=3).value = expense.amount
        row += 1

    output = BytesIO()

    workbook.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition":
                f"attachment; filename=budget_report_{project_id}.xlsx"
        }
    )