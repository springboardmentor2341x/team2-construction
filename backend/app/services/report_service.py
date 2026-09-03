import os
from io import BytesIO
from datetime import datetime, date

# ReportLab imports
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, KeepTogether, HRFlowable
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

# OpenPyXL imports
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ==============================================================================
# REPORTLAB STYLING HELPERS
# ==============================================================================

def get_pdf_styles():
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=20,
        leading=24,
        textColor=colors.HexColor('#1F4E78'),
        alignment=0,
        spaceAfter=10
    )
    
    subtitle_style = ParagraphStyle(
        'DocSubTitle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        leading=14,
        textColor=colors.HexColor('#555555'),
        spaceAfter=15
    )

    section_heading = ParagraphStyle(
        'SectionHeading',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=13,
        leading=17,
        textColor=colors.HexColor('#1F4E78'),
        spaceBefore=12,
        spaceAfter=6
    )

    cell_style = ParagraphStyle(
        'GridCell',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        leading=11,
        textColor=colors.HexColor('#222222')
    )

    cell_bold = ParagraphStyle(
        'GridCellBold',
        parent=cell_style,
        fontName='Helvetica-Bold'
    )

    header_cell = ParagraphStyle(
        'HeaderCell',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=12,
        textColor=colors.white
    )

    empty_cell = ParagraphStyle(
        'EmptyCell',
        parent=cell_style,
        fontName='Helvetica-Oblique',
        textColor=colors.HexColor('#777777')
    )

    return {
        'title': title_style,
        'subtitle': subtitle_style,
        'section': section_heading,
        'cell': cell_style,
        'cell_bold': cell_bold,
        'header': header_cell,
        'empty': empty_cell
    }


def create_meta_table(meta_data: list, col_widths=None):
    """Renders project info metadata block."""
    if not col_widths:
        col_widths = [120, 140, 120, 140]
    
    ps = get_pdf_styles()
    formatted_grid = []
    for row in meta_data:
        formatted_row = []
        for i, val in enumerate(row):
            if i % 2 == 0:
                formatted_row.append(Paragraph(str(val), ps['cell_bold']))
            else:
                formatted_row.append(Paragraph(str(val), ps['cell']))
        formatted_grid.append(formatted_row)
        
    t = Table(formatted_grid, colWidths=col_widths)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#F8FAFC')),
        ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#CBD5E1')),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
        ('PADDING', (0,0), (-1,-1), 5),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))
    return t


def create_report_table(headers: list, data: list, col_widths: list, empty_msg: str = "No records found."):
    """Renders a styled data table with headers and data rows."""
    ps = get_pdf_styles()
    table_rows = []
    
    # Header Row
    header_row = [Paragraph(str(h), ps['header']) for h in headers]
    table_rows.append(header_row)
    
    if not data:
        empty_row = [Paragraph(empty_msg, ps['empty'])] + [''] * (len(headers) - 1)
        table_rows.append(empty_row)
        t = Table(table_rows, colWidths=col_widths)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
            ('SPAN', (0, 1), (-1, 1)),
            ('ALIGN', (0, 1), (-1, 1), 'CENTER'),
            ('PADDING', (0, 0), (-1, -1), 5),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ]))
        return t
        
    for row_idx, row in enumerate(data):
        formatted_row = []
        for item in row:
            formatted_row.append(Paragraph(str(item if item is not None else '-'), ps['cell']))
        table_rows.append(formatted_row)
        
    t = Table(table_rows, colWidths=col_widths)
    t_style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
        ('PADDING', (0, 0), (-1, -1), 5),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]
    
    # Alternating row background
    for r in range(1, len(table_rows)):
        if r % 2 == 0:
            t_style.append(('BACKGROUND', (0, r), (-1, r), colors.HexColor('#F8FAFC')))
            
    t.setStyle(TableStyle(t_style))
    return t


# ==============================================================================
# OPENPYXL STYLING HELPERS
# ==============================================================================

def setup_excel_sheet(ws, title: str, meta_data: list, headers: list, data: list):
    """Pops headers, metadata, data and auto-fits columns in OpenPyXL sheet."""
    # Palette
    NAVY = "1F4E78"
    STEEL = "2E75B6"
    LIGHT_GRAY = "F2F4F7"
    BORDER_COLOR = "D9D9D9"

    font_title = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    font_sub = Font(name="Calibri", size=11, bold=True, color=NAVY)
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_data = Font(name="Calibri", size=10)
    font_empty = Font(name="Calibri", size=10, italic=True, color="777777")
    
    fill_title = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    fill_header = PatternFill(start_color=STEEL, end_color=STEEL, fill_type="solid")
    fill_alt = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")

    thin = Side(border_style="thin", color=BORDER_COLOR)
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title Banner
    ws.merge_cells("A1:G1")
    cell_title = ws["A1"]
    cell_title.value = title.upper()
    cell_title.font = font_title
    cell_title.fill = fill_title
    cell_title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # Metadata Block
    curr_row = 3
    for label, val in meta_data:
        ws.cell(row=curr_row, column=1, value=label).font = font_sub
        ws.cell(row=curr_row, column=2, value=val).font = font_data
        curr_row += 1

    curr_row += 1

    # Header Row
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=curr_row, column=col_idx, value=h)
        c.font = font_header
        c.fill = fill_header
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = cell_border
    ws.row_dimensions[curr_row].height = 24
    header_row_idx = curr_row
    curr_row += 1

    # Data Rows
    if not data:
        c = ws.cell(row=curr_row, column=1, value="No records found.")
        c.font = font_empty
        curr_row += 1
    else:
        for r_idx, row in enumerate(data):
            ws.row_dimensions[curr_row].height = 20
            for c_idx, val in enumerate(row, start=1):
                c = ws.cell(row=curr_row, column=c_idx, value=val if val is not None else "-")
                c.font = font_data
                c.border = cell_border
                if r_idx % 2 == 1:
                    c.fill = fill_alt
                # Alignment
                if isinstance(val, (int, float)):
                    c.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    c.alignment = Alignment(horizontal="left", vertical="center")
            curr_row += 1

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row == 1: # skip main banner
                continue
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 14)


# ==============================================================================
# 1. PROJECT PROGRESS REPORT EXPORTS
# ==============================================================================

def generate_project_progress_pdf(data: dict) -> BytesIO:
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=36,
        rightMargin=36,
        topMargin=36,
        bottomMargin=36
    )
    story = []
    ps = get_pdf_styles()

    # Header
    story.append(Paragraph("PROJECT PROGRESS REPORT", ps['title']))
    gen_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    story.append(Paragraph(f"Generated on: {gen_time} | BuildTrack System", ps['subtitle']))
    story.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor('#1F4E78'), spaceAfter=15))

    # Meta Table
    meta_info = [
        ["Project Name:", data.get('project_name'), "Project Code:", data.get('project_code')],
        ["Project Status:", data.get('project_status'), "Overall Progress:", f"{data.get('overall_progress')}%"],
        ["Start Date:", str(data.get('start_date') or '-'), "Target Completion:", str(data.get('expected_completion_date') or '-')],
        ["Project Manager:", data.get('project_manager') or 'Unassigned', "Project ID:", str(data.get('project_id'))]
    ]
    story.append(create_meta_table(meta_info))
    story.append(Spacer(1, 15))

    # 1. Milestones
    story.append(Paragraph("1. Project Milestones", ps['section']))
    m_headers = ["Title", "Planned Date", "Actual Completion", "Progress (%)", "Status"]
    m_data = [
        [m.get('title'), m.get('planned_date'), m.get('actual_completion_date') or '-', f"{m.get('progress_percentage')}%", m.get('status')]
        for m in data.get('milestones', [])
    ]
    story.append(create_report_table(m_headers, m_data, [150, 85, 95, 90, 100], "No milestones recorded."))
    story.append(Spacer(1, 15))

    # 2. Progress Updates
    story.append(Paragraph("2. Recent Site Progress Updates", ps['section']))
    pu_headers = ["Activity Name", "Update Date", "Progress (%)", "Status", "Updated By"]
    pu_data = [
        [pu.get('activity_name'), pu.get('update_date'), f"{pu.get('progress_percentage')}%", pu.get('status'), pu.get('updated_by') or '-']
        for pu in data.get('progress_updates', [])
    ]
    story.append(create_report_table(pu_headers, pu_data, [140, 90, 90, 95, 105], "No progress updates recorded."))
    story.append(Spacer(1, 15))

    # 3. Delays
    story.append(Paragraph("3. Delay Records", ps['section']))
    d_headers = ["Delay Date", "Reason", "Duration (Hrs)", "Work Category", "Status"]
    d_data = [
        [d.get('delay_date'), d.get('reason'), d.get('duration_hours'), d.get('affected_work_category') or '-', d.get('status')]
        for d in data.get('delays', [])
    ]
    story.append(create_report_table(d_headers, d_data, [85, 175, 80, 100, 80], "No delays recorded."))

    doc.build(story)
    buffer.seek(0)
    return buffer


def generate_project_progress_excel(data: dict) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Progress Overview"

    meta_list = [
        ("Project Name:", data.get('project_name')),
        ("Project Code:", data.get('project_code')),
        ("Status:", data.get('project_status')),
        ("Overall Progress:", f"{data.get('overall_progress')}%"),
        ("Project Manager:", data.get('project_manager') or 'Unassigned')
    ]
    headers = ["Type", "Name / Title / Reason", "Date", "Progress (%)", "Status", "Details / Remarks"]
    
    rows = []
    for m in data.get('milestones', []):
        rows.append(["Milestone", m.get('title'), m.get('planned_date'), m.get('progress_percentage'), m.get('status'), f"Completed: {m.get('actual_completion_date') or 'N/A'}"])
    for pu in data.get('progress_updates', []):
        rows.append(["Site Update", pu.get('activity_name'), pu.get('update_date'), pu.get('progress_percentage'), pu.get('status'), f"By: {pu.get('updated_by') or 'N/A'}"])
    for d in data.get('delays', []):
        rows.append(["Delay Record", d.get('reason'), d.get('delay_date'), 0, d.get('status'), f"Hours: {d.get('duration_hours')} | Category: {d.get('affected_work_category')}"])

    setup_excel_sheet(ws, "Project Progress Report", meta_list, headers, rows)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ==============================================================================
# 2. RESOURCE UTILIZATION REPORT EXPORTS
# ==============================================================================

def generate_resource_utilization_pdf(data: dict) -> BytesIO:
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=36)
    story = []
    ps = get_pdf_styles()

    story.append(Paragraph("RESOURCE UTILIZATION REPORT", ps['title']))
    gen_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    story.append(Paragraph(f"Generated on: {gen_time} | BuildTrack System", ps['subtitle']))
    story.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor('#1F4E78'), spaceAfter=15))

    meta_info = [
        ["Project Name:", data.get('project_name'), "Total Allocated Equipment:", str(data.get('total_allocated_equipment'))],
        ["Total Operating Hrs:", f"{data.get('total_operating_hours')} hrs", "Total Idle Hrs:", f"{data.get('total_idle_hours')} hrs"],
        ["Project ID:", str(data.get('project_id')), "Report Scope:", "Equipment & Materials"]
    ]
    story.append(create_meta_table(meta_info))
    story.append(Spacer(1, 15))

    # Equipment Allocations
    story.append(Paragraph("1. Equipment Allocations", ps['section']))
    eq_headers = ["Equipment Name", "Category", "Model / Serial", "Start Date", "End Date", "Status"]
    eq_data = [
        [
            ea.get('equipment_name'),
            ea.get('category'),
            f"{ea.get('model_number') or '-'} / {ea.get('serial_number') or '-'}",
            ea.get('start_date'),
            ea.get('end_date') or 'Ongoing',
            ea.get('status')
        ]
        for ea in data.get('equipment_allocations', [])
    ]
    story.append(create_report_table(eq_headers, eq_data, [120, 90, 110, 75, 75, 50], "No equipment allocations found."))
    story.append(Spacer(1, 15))

    # Equipment Utilization Logs
    story.append(Paragraph("2. Equipment Operating vs Idle Hours Log", ps['section']))
    u_headers = ["Equipment Name", "Usage Date", "Operating Hrs", "Idle Hrs", "Remarks"]
    u_data = [
        [eu.get('equipment_name'), eu.get('usage_date'), eu.get('operating_hours'), eu.get('idle_hours'), eu.get('remarks') or '-']
        for eu in data.get('equipment_utilization', [])
    ]
    story.append(create_report_table(u_headers, u_data, [130, 95, 95, 80, 120], "No utilization logs recorded."))
    story.append(Spacer(1, 15))

    # Material Allocations
    story.append(Paragraph("3. Material Allocations & Usage", ps['section']))
    mat_headers = ["Item Name", "Allocated Qty", "Allocation Date", "Work Activity", "Responsible"]
    mat_data = [
        [ma.get('item_name'), f"{ma.get('allocated_quantity')} {ma.get('unit') or ''}", ma.get('allocation_date'), ma.get('work_activity') or '-', ma.get('responsible_user') or '-']
        for ma in data.get('material_allocations', [])
    ]
    story.append(create_report_table(mat_headers, mat_data, [120, 90, 95, 115, 100], "No material allocations recorded."))

    doc.build(story)
    buffer.seek(0)
    return buffer


def generate_resource_utilization_excel(data: dict) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Equipment & Resources"

    meta_list = [
        ("Project Name:", data.get('project_name')),
        ("Total Equipment Allocated:", data.get('total_allocated_equipment')),
        ("Total Operating Hours:", f"{data.get('total_operating_hours')} hrs"),
        ("Total Idle Hours:", f"{data.get('total_idle_hours')} hrs")
    ]
    headers = ["Resource Type", "Item Name", "Category / Unit", "Quantity / Operating Hrs", "Idle Hrs / End Date", "Status / Remarks"]
    
    rows = []
    for ea in data.get('equipment_allocations', []):
        rows.append(["Equipment Allocation", ea.get('equipment_name'), ea.get('category'), 1, f"Start: {ea.get('start_date')} | End: {ea.get('end_date') or 'Ongoing'}", ea.get('status')])
    for eu in data.get('equipment_utilization', []):
        rows.append(["Equipment Utilization Log", eu.get('equipment_name'), f"Date: {eu.get('usage_date')}", f"Operating: {eu.get('operating_hours')} hrs", f"Idle: {eu.get('idle_hours')} hrs", eu.get('remarks') or '-'])
    for ma in data.get('material_allocations', []):
        rows.append(["Material Allocation", ma.get('item_name'), ma.get('unit'), ma.get('allocated_quantity'), ma.get('allocation_date'), f"Activity: {ma.get('work_activity')}"])

    setup_excel_sheet(ws, "Resource Utilization Report", meta_list, headers, rows)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ==============================================================================
# 3. WORKFORCE REPORT EXPORTS
# ==============================================================================

def generate_workforce_pdf(data: dict) -> BytesIO:
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=36)
    story = []
    ps = get_pdf_styles()

    story.append(Paragraph("WORKFORCE MANAGEMENT REPORT", ps['title']))
    gen_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    story.append(Paragraph(f"Generated on: {gen_time} | BuildTrack System", ps['subtitle']))
    story.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor('#1F4E78'), spaceAfter=15))

    att_summary = data.get('attendance_summary', {})
    meta_info = [
        ["Project Name:", data.get('project_name'), "Total Assigned Workers:", str(data.get('total_assigned_workers'))],
        ["Present Count:", str(att_summary.get('total_present', 0)), "Absent Count:", str(att_summary.get('total_absent', 0))],
        ["Leave Count:", str(att_summary.get('total_leave', 0)), "Total Working Hours:", f"{att_summary.get('total_working_hours', 0)} hrs"]
    ]
    story.append(create_meta_table(meta_info))
    story.append(Spacer(1, 15))

    # Assigned Workers
    story.append(Paragraph("1. Worker Assignments", ps['section']))
    w_headers = ["Worker Name", "Designation", "Work Activity", "Contractor", "Start Date", "Status"]
    w_data = [
        [wa.get('worker_name'), wa.get('designation') or '-', wa.get('work_activity') or '-', wa.get('contractor_name') or '-', wa.get('start_date'), wa.get('assignment_status')]
        for wa in data.get('worker_assignments', [])
    ]
    story.append(create_report_table(w_headers, w_data, [110, 95, 105, 95, 75, 40], "No worker assignments found."))
    story.append(Spacer(1, 15))

    # Attendance Records
    story.append(Paragraph("2. Daily Attendance Records", ps['section']))
    att_headers = ["Worker Name", "Date", "Status", "Check In", "Check Out", "Hours"]
    att_data = [
        [a.get('worker_name'), a.get('date'), a.get('status'), a.get('check_in_time') or '-', a.get('check_out_time') or '-', a.get('working_hours') or 0]
        for a in data.get('attendance_records', [])
    ]
    story.append(create_report_table(att_headers, att_data, [140, 90, 80, 75, 75, 60], "No attendance records found."))
    story.append(Spacer(1, 15))

    # Payroll Summary
    story.append(Paragraph("3. Payroll & Wage Summary", ps['section']))
    p_headers = ["Worker Name", "Pay Rate", "Work Days", "Work Hours", "Estimated Pay", "Payroll Status"]
    p_data = [
        [p.get('worker_name'), f"${p.get('pay_rate')}", p.get('working_days'), p.get('working_hours'), f"${p.get('estimated_pay')}", p.get('payroll_status')]
        for p in data.get('payroll', [])
    ]
    story.append(create_report_table(p_headers, p_data, [130, 75, 70, 75, 90, 80], "No payroll records found."))

    doc.build(story)
    buffer.seek(0)
    return buffer


def generate_workforce_excel(data: dict) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Workforce & Attendance"

    att_summary = data.get('attendance_summary', {})
    meta_list = [
        ("Project Name:", data.get('project_name')),
        ("Total Assigned Workers:", data.get('total_assigned_workers')),
        ("Present Days Count:", att_summary.get('total_present', 0)),
        ("Absent Days Count:", att_summary.get('total_absent', 0)),
        ("Leave Days Count:", att_summary.get('total_leave', 0))
    ]
    headers = ["Category", "Worker Name", "Designation / Date", "Work Activity / Status", "Hours / Pay Rate", "Est. Pay / Details"]
    
    rows = []
    for wa in data.get('worker_assignments', []):
        rows.append(["Worker Assignment", wa.get('worker_name'), wa.get('designation'), wa.get('work_activity'), f"Start: {wa.get('start_date')}", f"Status: {wa.get('assignment_status')}"])
    for a in data.get('attendance_records', []):
        rows.append(["Attendance Log", a.get('worker_name'), str(a.get('date')), a.get('status'), f"{a.get('working_hours')} hrs", f"In: {a.get('check_in_time')} | Out: {a.get('check_out_time')}"])
    for p in data.get('payroll', []):
        rows.append(["Payroll Record", p.get('worker_name'), f"Days: {p.get('working_days')}", f"Rate: ${p.get('pay_rate')}", f"Hours: {p.get('working_hours')}", f"Pay: ${p.get('estimated_pay')} ({p.get('payroll_status')})"])

    setup_excel_sheet(ws, "Workforce Management Report", meta_list, headers, rows)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ==============================================================================
# 4. PROCUREMENT REPORT EXPORTS
# ==============================================================================

def generate_procurement_pdf(data: dict) -> BytesIO:
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=36)
    story = []
    ps = get_pdf_styles()

    story.append(Paragraph("PROCUREMENT & VENDOR REPORT", ps['title']))
    gen_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    story.append(Paragraph(f"Generated on: {gen_time} | BuildTrack System", ps['subtitle']))
    story.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor('#1F4E78'), spaceAfter=15))

    meta_info = [
        ["Project Name:", data.get('project_name'), "Total Purchase Orders:", f"${data.get('total_po_value'):,.2f}"],
        ["Total Invoiced:", f"${data.get('total_invoiced_amount'):,.2f}", "Total Paid Amount:", f"${data.get('total_paid_amount'):,.2f}"],
        ["Outstanding Amount:", f"${data.get('outstanding_amount'):,.2f}", "Project ID:", str(data.get('project_id'))]
    ]
    story.append(create_meta_table(meta_info))
    story.append(Spacer(1, 15))

    # Procurement Requests
    story.append(Paragraph("1. Material & Item Procurement Requests", ps['section']))
    req_headers = ["Item Name", "Category", "Req Qty", "Required Date", "Requested By", "Status"]
    req_data = [
        [pr.get('item_name'), pr.get('category') or '-', pr.get('requested_quantity'), pr.get('required_date') or '-', pr.get('requested_by'), pr.get('request_status')]
        for pr in data.get('procurement_requests', [])
    ]
    story.append(create_report_table(req_headers, req_data, [120, 85, 60, 85, 100, 70], "No procurement requests found."))
    story.append(Spacer(1, 15))

    # Purchase Orders
    story.append(Paragraph("2. Vendor Purchase Orders", ps['section']))
    po_headers = ["PO ID", "Vendor Name", "Order Date", "Expected Date", "Total Amount", "Status"]
    po_data = [
        [f"PO-{po.get('id')}", po.get('vendor_name'), po.get('order_date') or '-', po.get('expected_delivery_date') or '-', f"${po.get('total_amount'):,.2f}", po.get('order_status')]
        for po in data.get('purchase_orders', [])
    ]
    story.append(create_report_table(po_headers, po_data, [60, 150, 80, 85, 85, 60], "No purchase orders issued."))
    story.append(Spacer(1, 15))

    # Invoices
    story.append(Paragraph("3. Invoices & Payment Status", ps['section']))
    inv_headers = ["Invoice No", "Vendor Name", "Invoice Date", "Amount", "Payment Status", "Invoice Status"]
    inv_data = [
        [inv.get('invoice_number'), inv.get('vendor_name'), inv.get('invoice_date') or '-', f"${inv.get('invoice_amount'):,.2f}", inv.get('payment_status'), inv.get('invoice_status')]
        for inv in data.get('invoices', [])
    ]
    story.append(create_report_table(inv_headers, inv_data, [90, 140, 80, 80, 70, 60], "No invoices recorded."))

    doc.build(story)
    buffer.seek(0)
    return buffer


def generate_procurement_excel(data: dict) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Procurement Summary"

    meta_list = [
        ("Project Name:", data.get('project_name')),
        ("Total Purchase Order Amount:", f"${data.get('total_po_value'):,.2f}"),
        ("Total Invoiced Amount:", f"${data.get('total_invoiced_amount'):,.2f}"),
        ("Total Paid Amount:", f"${data.get('total_paid_amount'):,.2f}"),
        ("Outstanding Balance:", f"${data.get('outstanding_amount'):,.2f}")
    ]
    headers = ["Record Type", "Item / Vendor / Inv No", "Category / Dates", "Quantity / PO Amt", "Invoiced / Paid Amt", "Status"]
    
    rows = []
    for pr in data.get('procurement_requests', []):
        rows.append(["Procurement Request", pr.get('item_name'), f"Req Date: {pr.get('required_date')}", pr.get('requested_quantity'), "-", pr.get('request_status')])
    for po in data.get('purchase_orders', []):
        rows.append(["Purchase Order", po.get('vendor_name'), f"Order Date: {po.get('order_date')}", f"${po.get('total_amount'):,.2f}", "-", po.get('order_status')])
    for inv in data.get('invoices', []):
        rows.append(["Invoice Record", f"{inv.get('invoice_number')} ({inv.get('vendor_name')})", f"Inv Date: {inv.get('invoice_date')}", "-", f"${inv.get('invoice_amount'):,.2f}", f"Pay: {inv.get('payment_status')}"])

    setup_excel_sheet(ws, "Procurement & Vendor Report", meta_list, headers, rows)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ==============================================================================
# 5. BUDGET REPORT EXPORTS
# ==============================================================================

def generate_budget_pdf(data: dict) -> BytesIO:
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=36)
    story = []
    ps = get_pdf_styles()

    story.append(Paragraph("PROJECT BUDGET & FINANCIAL REPORT", ps['title']))
    gen_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    story.append(Paragraph(f"Generated on: {gen_time} | BuildTrack System", ps['subtitle']))
    story.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor('#1F4E78'), spaceAfter=15))

    meta_info = [
        ["Project Name:", data.get('project_name'), "Total Approved Budget:", f"${data.get('total_budget'):,.2f}"],
        ["Total Estimated Cost:", f"${data.get('total_estimated_cost'):,.2f}", "Total Actual Expenses:", f"${data.get('total_actual_expenses'):,.2f}"],
        ["Remaining Budget:", f"${data.get('remaining_budget'):,.2f}", "Budget Utilized (%):", f"{data.get('utilization_percentage')}%"]
    ]
    story.append(create_meta_table(meta_info))
    story.append(Spacer(1, 15))

    # Budget Allocation Breakdown by Category
    story.append(Paragraph("1. Planned Budget Allocation Breakdown", ps['section']))
    cat_headers = ["Budget Category", "Allocated Amount", "Actual Spent", "Remaining Balance"]
    cat_data = [
        [v.get('category'), f"${v.get('allocated'):,.2f}", f"${v.get('actual_spent'):,.2f}", f"${v.get('remaining'):,.2f}"]
        for v in data.get('category_variance', [])
    ]
    story.append(create_report_table(cat_headers, cat_data, [150, 120, 125, 125], "No category breakdowns available."))
    story.append(Spacer(1, 15))

    # Cost Estimates
    story.append(Paragraph("2. Cost Estimates", ps['section']))
    est_headers = ["Category", "Description", "Estimate Date", "Estimated Amount", "Status"]
    est_data = [
        [e.get('category'), e.get('description') or '-', e.get('estimate_date') or '-', f"${e.get('estimated_amount'):,.2f}", e.get('status')]
        for e in data.get('estimates', [])
    ]
    story.append(create_report_table(est_headers, est_data, [110, 170, 80, 90, 70], "No cost estimates recorded."))
    story.append(Spacer(1, 15))

    # Actual Expenses Log
    story.append(Paragraph("3. Recorded Expenses Log", ps['section']))
    exp_headers = ["Category", "Description", "Expense Date", "Amount", "Recorded By"]
    exp_data = [
        [e.get('category'), e.get('description') or '-', e.get('expense_date') or '-', f"${e.get('amount'):,.2f}", e.get('recorded_by') or '-']
        for e in data.get('expenses', [])
    ]
    story.append(create_report_table(exp_headers, exp_data, [110, 170, 80, 85, 75], "No actual expenses recorded."))

    doc.build(story)
    buffer.seek(0)
    return buffer


def generate_budget_excel(data: dict) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Budget Overview"

    meta_list = [
        ("Project Name:", data.get('project_name')),
        ("Total Approved Budget:", f"${data.get('total_budget'):,.2f}"),
        ("Total Estimated Cost:", f"${data.get('total_estimated_cost'):,.2f}"),
        ("Total Actual Expenses:", f"${data.get('total_actual_expenses'):,.2f}"),
        ("Remaining Budget:", f"${data.get('remaining_budget'):,.2f}"),
        ("Utilization Percentage:", f"{data.get('utilization_percentage')}%")
    ]
    headers = ["Type", "Category", "Description / Details", "Date", "Estimated / Allocated", "Actual Expense"]
    
    rows = []
    for cv in data.get('category_variance', []):
        rows.append(["Category Allocation", cv.get('category'), "Allocated vs Actual Spent", "-", f"${cv.get('allocated'):,.2f}", f"${cv.get('actual_spent'):,.2f}"])
    for est in data.get('estimates', []):
        rows.append(["Cost Estimate", est.get('category'), est.get('description'), est.get('estimate_date'), f"${est.get('estimated_amount'):,.2f}", "-"])
    for exp in data.get('expenses', []):
        rows.append(["Actual Expense", exp.get('category'), exp.get('description'), exp.get('expense_date'), "-", f"${exp.get('amount'):,.2f}"])

    setup_excel_sheet(ws, "Project Budget & Cost Management Report", meta_list, headers, rows)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
